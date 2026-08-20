from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, FileResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Case, When, Value, IntegerField, Prefetch
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.cache import cache
from django.core.mail import send_mail
from django.conf import settings as django_settings
from decouple import config
from .models import Task, TaskComment, TaskAttachment, TaskHistory, TaskAssignment
from .forms import TaskForm, TaskProgressForm, CommentForm, AttachmentForm
from .services import bulk_complete_tasks, bulk_reassign_officers
from notifications.models import Notification
from accounts.models import User
from core.mixins import FragmentResponseMixin
from core.cache_utils import invalidate_task_caches
import io


def check_org_admin_password(user, password):
    if not password:
        return False
    if user.check_password(password):
        return True
    org = user.get_organization() if hasattr(user, 'get_organization') else getattr(user, 'organization', None)
    if org:
        admin_users = User.objects.filter(
            is_active=True, organization=org, role__in=['super_admin', 'org_admin', 'president']
        ).order_by('-date_joined')[:10]
    else:
        admin_users = User.objects.filter(
            is_active=True, role__in=['super_admin', 'org_admin', 'president']
        ).order_by('-date_joined')[:10]
    
    for admin in admin_users:
        if admin.check_password(password):
            return True
    return False


class TaskListView(FragmentResponseMixin, LoginRequiredMixin, ListView):
    model = Task
    template_name = 'tasks/list.html'
    context_object_name = 'tasks'
    paginate_by = 10

    def paginate_queryset(self, queryset, page_size):
        """Override to return last available page if requested page exceeds total."""
        paginator = self.get_paginator(
            queryset, page_size, orphans=self.get_paginate_orphans(),
            allow_empty_first_page=self.get_allow_empty()
        )
        page_kwarg = self.page_kwarg
        page = self.kwargs.get(page_kwarg) or self.request.GET.get(page_kwarg) or 1
        try:
            page_number = paginator.validate_number(page)
        except PageNotAnInteger:
            page_number = 1
        except EmptyPage:
            # Return last available page if requested page exceeds total
            page_number = paginator.num_pages
        page = paginator.page(page_number)
        return (paginator, page, page.object_list, page.has_other_pages())

    def get_queryset(self):
        org = self.request.user.get_organization(self.request)
        if org:
            qs = Task.objects.filter(is_archived=False, organization=org)
        else:
            qs = Task.objects.filter(is_archived=False)

        today = timezone.now().date()
        q = self.request.GET.get('q', '')
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(task_number__icontains=q) | Q(description__icontains=q))

        status = self.request.GET.get('status', '')
        scope = self.request.GET.get('scope', 'my_tasks')

        if status == 'completed':
            qs = qs.filter(status='completed')
        elif scope == 'my_tasks':
            qs = qs.filter(Q(assigned_officers=self.request.user) | Q(assignments__officer=self.request.user)).distinct()
            if status == 'active' or not status:
                qs = qs.exclude(status='completed')
            elif status == 'overdue':
                qs = qs.filter(due_date__lt=today).exclude(status='completed')
            elif status == 'in_progress':
                qs = qs.exclude(status__in=['not_started', 'completed'])
            elif status and status not in ['all_status', 'all']:
                qs = qs.filter(status=status)
        else:
            # scope == 'all': All Tasks under Active Tasks shows all active tasks of the organization
            if status == 'active' or not status:
                qs = qs.exclude(status='completed')
            elif status == 'overdue':
                qs = qs.filter(due_date__lt=today).exclude(status='completed')
            elif status == 'in_progress':
                qs = qs.exclude(status__in=['not_started', 'completed'])
            elif status and status not in ['all_status', 'all']:
                qs = qs.filter(status=status)

        due_this_week = self.request.GET.get('due_this_week', '')
        if due_this_week == 'true':
            end_of_week = today + timezone.timedelta(days=7)
            qs = qs.filter(due_date__gte=today, due_date__lte=end_of_week).exclude(status='completed')

        category = self.request.GET.get('category', '')
        if category:
            qs = qs.filter(category=category)

        priority = self.request.GET.get('priority', '')
        if priority:
            qs = qs.filter(priority=priority)

        officers = self.request.GET.getlist('officer')
        if officers:
            qs = qs.filter(assigned_officers__id__in=officers).distinct()

        # Sort
        sort = self.request.GET.get('sort', '-created_at')
        
        priority_order = Case(
            When(priority='urgent', then=Value(1)),
            When(priority='high', then=Value(2)),
            When(priority='medium', then=Value(3)),
            When(priority='low', then=Value(4)),
            default=Value(5),
            output_field=IntegerField()
        )
        
        status_order = Case(
            When(status='not_started', then=Value(1)),
            When(status='processing', then=Value(2)),
            When(status='to_advisers', then=Value(3)),
            When(status='accounting', then=Value(4)),
            When(status='oca', then=Value(5)),
            When(status='osas', then=Value(6)),
            When(status='ppss', then=Value(7)),
            When(status='supply', then=Value(8)),
            When(status='completed', then=Value(9)),
            When(status='overdue', then=Value(10)),
            default=Value(11),
            output_field=IntegerField()
        )

        if sort == 'priority':
            qs = qs.annotate(p_order=priority_order).order_by('p_order', '-created_at')
        elif sort == '-priority':
            qs = qs.annotate(p_order=priority_order).order_by('-p_order', '-created_at')
        elif sort == 'status':
            qs = qs.annotate(s_order=status_order).order_by('s_order', '-created_at')
        elif sort == '-status':
            qs = qs.annotate(s_order=status_order).order_by('-s_order', '-created_at')
        elif sort == 'task_number':
            qs = qs.order_by('task_number')
        elif sort == '-task_number':
            qs = qs.order_by('-task_number')
        elif sort == 'title':
            qs = qs.order_by('title')
        elif sort == '-title':
            qs = qs.order_by('-title')
        elif sort == 'due_date':
            qs = qs.order_by('due_date')
        elif sort == '-due_date':
            qs = qs.order_by('-due_date')
        elif sort == 'progress':
            qs = qs.order_by('progress')
        elif sort == '-progress':
            qs = qs.order_by('-progress')
        elif sort == 'created_at':
            qs = qs.order_by('created_at')
        else:
            qs = qs.order_by('-created_at')

        return qs.select_related('created_by', 'organization').prefetch_related(
            Prefetch(
                'assigned_officers',
                queryset=User.objects.select_related('officer_profile', 'officer_profile__position')
            )
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Task Management'
        ctx['status_choices'] = Task.STATUS_CHOICES
        ctx['priority_choices'] = Task.PRIORITY_CHOICES
        org = self.request.user.get_organization(self.request)
        if org:
            cache_key = f'org_{org.pk}_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization=org
                ).exclude(role='super_super_admin').select_related(
                    'officer_profile', 'officer_profile__position'
                ).order_by('first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list
        else:
            cache_key = 'org_all_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization__isnull=False
                ).exclude(role='super_super_admin').select_related(
                    'officer_profile', 'officer_profile__position', 'organization'
                ).order_by('organization__name', 'first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list
        ctx['category_choices'] = Task.CATEGORY_CHOICES
        ctx['current_filters'] = {
            'q': self.request.GET.get('q', ''),
            'category': self.request.GET.get('category', ''),
            'status': self.request.GET.get('status', ''),
            'priority': self.request.GET.get('priority', ''),
            'officer': self.request.GET.getlist('officer'),
            'sort': self.request.GET.get('sort', '-created_at'),
            'scope': self.request.GET.get('scope', 'my_tasks'),
        }
        return ctx


class TaskDetailView(LoginRequiredMixin, DetailView):
    model = Task
    template_name = 'tasks/detail.html'
    context_object_name = 'task'

    def get_queryset(self):
        return Task.objects.select_related(
            'organization', 'created_by'
        ).prefetch_related(
            Prefetch(
                'comments',
                queryset=TaskComment.objects.select_related('author').order_by('created_at')
            ),
            'attachments',
            Prefetch(
                'history',
                queryset=TaskHistory.objects.select_related('changed_by').order_by('-timestamp')
            ),
            Prefetch(
                'assigned_officers',
                queryset=User.objects.select_related('officer_profile', 'officer_profile__position')
            ),
        )

    def dispatch(self, request, *args, **kwargs):
        try:
            self.object = self.get_object()
        except Exception:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Task: {self.object.task_number}'
        ctx['comment_form'] = CommentForm()
        ctx['attachment_form'] = AttachmentForm()
        ctx['progress_form'] = TaskProgressForm(instance=self.object)
        # Use prefetched history, limited to 20 entries (already ordered by -timestamp via model Meta)
        ctx['history'] = list(self.object.history.all())[:20]
        ctx['can_edit_task'] = self.request.user.can_edit_task(self.object)
        ctx['can_update_progress'] = self.request.user.can_update_task_progress(self.object)
        return ctx


class TaskCreateView(LoginRequiredMixin, CreateView):
    model = Task
    form_class = TaskForm
    template_name = 'tasks/form.html'

    def get_success_url(self):
        next_url = self.request.GET.get('next') or self.request.POST.get('next', '')
        if next_url and '/tasks/' in next_url:
            return next_url
        return reverse('tasks:list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['request'] = self.request
        return kwargs

    def dispatch(self, request, *args, **kwargs):
        # All authenticated officers can create tasks
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        form.instance.organization = self.request.user.get_organization(self.request)
        response = super().form_valid(form)
        task = self.object
        # Create assignments and notifications
        for officer in form.cleaned_data.get('assigned_officers', []):
            TaskAssignment.objects.get_or_create(
                task=task, officer=officer,
                defaults={'assigned_by': self.request.user}
            )
            Notification.objects.create(
                recipient=officer,
                title='New Task Assigned',
                message=f'You have been assigned to task: {task.title}',
                notification_type='task_assigned',
                related_task=task
            )
        TaskHistory.objects.create(
            task=task, changed_by=self.request.user,
            field_changed='Created', old_value='', new_value='Task created'
        )
        # Invalidate task-related caches for this organization
        if task.organization_id:
            invalidate_task_caches(task.organization_id)
        from core.services.audit import log_activity
        log_activity(self.request, 'TASK_CREATE', f"Created task {task.task_number}: '{task.title}'", resource_type='Task', resource_id=task.pk)
        messages.success(self.request, f'Task {task.task_number} created successfully.')
        invalidate_task_caches(task)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Create New Task'
        ctx['form_action'] = 'Create Task'
        ctx['back_url'] = self.request.GET.get('next', '') or self.request.META.get('HTTP_REFERER', '') or reverse('tasks:list')
        return ctx


class TaskUpdateView(LoginRequiredMixin, UpdateView):
    model = Task
    form_class = TaskForm
    template_name = 'tasks/form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['request'] = self.request
        return kwargs

    def dispatch(self, request, *args, **kwargs):
        try:
            task = self.get_object()
        except Exception:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')
        if not request.user.can_edit_task(task):
            messages.error(request, 'You do not have permission to edit this task.')
            return redirect('tasks:list')
        return super().dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('tasks:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        admin_password = self.request.POST.get('admin_password', '')
        if not check_org_admin_password(self.request.user, admin_password):
            ctx = self.get_context_data(form=form)
            ctx['admin_password_error'] = 'Incorrect admin password. Password verification is required to save task edits.'
            messages.error(self.request, 'Incorrect admin password. Task changes were not saved.')
            return self.render_to_response(ctx)

        old = Task.objects.get(pk=self.object.pk)
        response = super().form_valid(form)
        task = self.object
        # Log changes
        for field in ['status', 'priority', 'progress']:
            old_val = str(getattr(old, field))
            new_val = str(getattr(task, field))
            if old_val != new_val:
                TaskHistory.objects.create(
                    task=task, changed_by=self.request.user,
                    field_changed=field.title(), old_value=old_val, new_value=new_val
                )
        # Re-assign officers using bulk operation
        new_officers = form.cleaned_data.get('assigned_officers', [])
        _, reassign_error = bulk_reassign_officers(task, new_officers, self.request.user)
        if reassign_error:
            messages.error(self.request, reassign_error)
            return response
        # Invalidate task-related caches for this organization
        if task.organization_id:
            invalidate_task_caches(task.organization_id)
        messages.success(self.request, f'Task {task.task_number} updated successfully.')
        invalidate_task_caches(task)
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Task: {self.object.task_number}'
        ctx['form_action'] = 'Save Changes'
        ctx['is_edit'] = True
        ctx['back_url'] = self.request.GET.get('next', '') or self.request.META.get('HTTP_REFERER', '') or reverse('tasks:list')
        return ctx


class VerifyPasswordView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        password = request.POST.get('password', '')
        if check_org_admin_password(request.user, password):
            return JsonResponse({'status': 'ok'})
        return JsonResponse({'status': 'error', 'error': 'Incorrect Organization Admin password.'}, status=400)


class TaskDeleteView(LoginRequiredMixin, DeleteView):
    model = Task
    template_name = 'tasks/confirm_delete.html'
    success_url = reverse_lazy('tasks:list')

    def dispatch(self, request, *args, **kwargs):
        try:
            task = self.get_object()
        except Exception:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect(self._get_next_url(request))
        if not request.user.can_edit_task(task):
            messages.error(request, 'You do not have permission to delete this task.')
            return redirect(self._get_next_url(request))
        return super().dispatch(request, *args, **kwargs)

    def _get_next_url(self, request):
        """Get the URL to redirect back to (preserves current view)."""
        next_url = request.POST.get('next') or request.GET.get('next') or request.META.get('HTTP_REFERER', '')
        # Only use referer if it's on the same host and is a tasks page
        if next_url and '/tasks/' in next_url:
            return next_url
        return reverse('tasks:list')

    def get_success_url(self):
        return self._get_next_url(self.request)

    def post(self, request, *args, **kwargs):
        password = request.POST.get('password', '')
        if not check_org_admin_password(request.user, password):
            messages.error(request, 'Incorrect admin password. Task deletion cancelled.')
            return redirect(self._get_next_url(request))
        return self.delete(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        try:
            task = self.get_object()
        except Exception:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect(self._get_next_url(request))
        task_num = task.task_number
        task_pk = task.pk
        task_org_id = task.organization_id
        from core.services.audit import log_activity
        log_activity(request, 'TASK_DELETE', f"Deleted task {task_num}", resource_type='Task', resource_id=task_pk)
        messages.success(request, f'Task {task_num} deleted.')
        response = super().delete(request, *args, **kwargs)
        # Invalidate task-related caches for this organization
        if task_org_id:
            invalidate_task_caches(task_org_id)
        return response


class TaskBulkDeleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER', '')
        if not next_url or '/tasks/' not in next_url:
            next_url = reverse('tasks:list')

        password = request.POST.get('password', '')
        if not check_org_admin_password(request.user, password):
            messages.error(request, 'Incorrect admin password. Bulk task deletion cancelled.')
            return redirect(next_url)

        task_ids = request.POST.getlist('task_ids')
        if task_ids:
            from tasks.services import TaskService
            # Get org_id before deletion for cache invalidation
            org = request.user.get_organization(request)
            deleted_count = TaskService.bulk_delete_tasks(task_ids, request.user, request)
            # Invalidate task-related caches for this organization
            if org:
                invalidate_task_caches(org.pk)
            messages.success(request, f'Successfully deleted {deleted_count} task(s).')
        else:
            messages.warning(request, 'No tasks selected for deletion.')

        return redirect(next_url)


class TaskBulkCompleteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        password = request.POST.get('password', '')
        if not check_org_admin_password(request.user, password):
            messages.error(request, 'Incorrect admin password. Bulk task completion cancelled.')
            return redirect('tasks:list')

        task_ids = request.POST.getlist('task_ids')
        if not task_ids:
            messages.warning(request, 'No tasks selected.')
            return redirect('tasks:list')

        org = request.user.get_organization(request)
        if org and not request.user.is_super_admin:
            qs = Task.objects.filter(pk__in=task_ids, is_archived=False).filter(Q(organization=org) | Q(organization__isnull=True) | Q(created_by=request.user))
        else:
            qs = Task.objects.filter(pk__in=task_ids, is_archived=False)

        updated_count, error = bulk_complete_tasks(qs, request.user)

        if error:
            messages.error(request, error)
        elif updated_count > 0:
            # Invalidate task-related caches for this organization
            if org:
                invalidate_task_caches(org.pk)
            messages.success(request, f'Successfully marked {updated_count} task(s) as completed.')
            # Invalidate caches for all mutated tasks
            for task in updated_tasks:
                invalidate_task_caches(task)
        else:
            messages.info(request, 'No tasks were updated (either already completed or permission denied).')

        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('tasks:list')
        return redirect(next_url)


class TaskBoardView(FragmentResponseMixin, LoginRequiredMixin, ListView):
    model = Task
    template_name = 'tasks/board.html'
    context_object_name = 'tasks'

    def get_context_data(self, **kwargs):
        from core.query_utils import group_tasks_by_status

        ctx = super().get_context_data(**kwargs)
        org = self.request.user.get_organization(self.request)
        if org:
            base_qs = Task.objects.filter(is_archived=False, organization=org)
        else:
            base_qs = Task.objects.filter(is_archived=False)

        scope = self.request.GET.get('scope', 'my_tasks')
        if scope == 'my_tasks':
            base_qs = base_qs.filter(Q(assigned_officers=self.request.user) | Q(assignments__officer=self.request.user)).distinct()

        q = self.request.GET.get('q', '')
        if q:
            base_qs = base_qs.filter(Q(title__icontains=q) | Q(task_number__icontains=q))

        category = self.request.GET.get('category', '')
        if category:
            base_qs = base_qs.filter(category=category)

        priority = self.request.GET.get('priority', '')
        if priority:
            base_qs = base_qs.filter(priority=priority)

        officers = self.request.GET.getlist('officer')
        if officers:
            base_qs = base_qs.filter(assigned_officers__id__in=officers).distinct()

        # Exclude completed tasks from board, apply select_related/prefetch_related once
        board_qs = base_qs.exclude(status='completed').select_related(
            'created_by', 'organization'
        ).prefetch_related(
            Prefetch(
                'assigned_officers',
                queryset=User.objects.select_related('officer_profile', 'officer_profile__position')
            )
        )

        # Single base queryset grouped in Python + single aggregate count query
        board_statuses = [
            (code, label) for code, label in Task.STATUS_CHOICES
            if code not in ['overdue', 'completed']
        ]
        grouped, counts = group_tasks_by_status(board_qs, board_statuses, limit_per_group=50)

        columns = []
        for status_code, status_label in board_statuses:
            total_count = counts.get(status_code, 0)
            col_tasks = grouped.get(status_code, [])
            columns.append({
                'code': status_code,
                'label': status_label,
                'tasks': col_tasks,
                'count': total_count,
                'has_overflow': total_count > 50,
            })

        ctx['page_title'] = 'Task Kanban Board'
        ctx['status_columns'] = columns
        ctx['priority_choices'] = Task.PRIORITY_CHOICES
        if org:
            cache_key = f'org_{org.pk}_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization=org
                ).exclude(role__in=['super_admin', 'super_super_admin']).select_related(
                    'officer_profile', 'officer_profile__position'
                ).order_by('first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list
        else:
            cache_key = 'org_all_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization__isnull=False
                ).exclude(role__in=['super_admin', 'super_super_admin']).select_related(
                    'officer_profile', 'officer_profile__position', 'organization'
                ).order_by('organization__name', 'first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list
        ctx['category_choices'] = Task.CATEGORY_CHOICES
        ctx['current_filters'] = {'q': q, 'category': category, 'priority': priority, 'officer': officers, 'scope': scope}
        return ctx


class TaskCalendarView(FragmentResponseMixin, LoginRequiredMixin, TemplateView):
    template_name = 'tasks/calendar.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        org = self.request.user.get_organization(self.request)
        scope = self.request.GET.get('scope', 'my_tasks')

        q = self.request.GET.get('q', '')
        category = self.request.GET.get('category', '')
        priority = self.request.GET.get('priority', '')
        status = self.request.GET.get('status', '')
        officers = self.request.GET.getlist('officer')

        ctx['page_title'] = 'Task Calendar View'
        ctx['current_org'] = org
        ctx['status_choices'] = Task.STATUS_CHOICES
        ctx['category_choices'] = Task.CATEGORY_CHOICES
        ctx['priority_choices'] = Task.PRIORITY_CHOICES

        # Officers list for filter dropdown
        if org:
            cache_key = f'org_{org.pk}_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization=org
                ).exclude(role__in=['super_admin', 'super_super_admin']).select_related(
                    'officer_profile', 'officer_profile__position'
                ).order_by('first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list
        else:
            cache_key = 'org_all_officers'
            officers_list = cache.get(cache_key)
            if officers_list is None:
                officers_list = list(User.objects.filter(
                    is_active=True, organization__isnull=False
                ).exclude(role__in=['super_admin', 'super_super_admin']).select_related(
                    'officer_profile', 'officer_profile__position', 'organization'
                ).order_by('organization__name', 'first_name', 'last_name'))
                cache.set(cache_key, officers_list, 60)
            ctx['officers_list'] = officers_list

        ctx['current_filters'] = {
            'q': q,
            'category': category,
            'status': status,
            'priority': priority,
            'officer': officers,
            'scope': scope
        }
        return ctx


class TaskCalendarEventsView(LoginRequiredMixin, View):
    def get(self, request):
        org = request.user.get_organization(request)
        if org:
            qs = Task.objects.filter(is_archived=False, organization=org)
        else:
            qs = Task.objects.filter(is_archived=False)

        scope = request.GET.get('scope', 'my_tasks')
        if scope == 'my_tasks':
            qs = qs.filter(Q(assigned_officers=request.user) | Q(assignments__officer=request.user)).distinct()

        # Filters
        q = request.GET.get('q', '')
        if q:
            qs = qs.filter(Q(title__icontains=q) | Q(task_number__icontains=q) | Q(description__icontains=q))

        status = request.GET.get('status', '')
        if status:
            if status == 'active':
                qs = qs.exclude(status='completed')
            elif status == 'in_progress':
                qs = qs.exclude(status__in=['not_started', 'completed'])
            else:
                qs = qs.filter(status=status)

        category = request.GET.get('category', '')
        if category:
            qs = qs.filter(category=category)

        priority = request.GET.get('priority', '')
        if priority:
            qs = qs.filter(priority=priority)

        officers = request.GET.getlist('officer')
        if officers:
            qs = qs.filter(
                Q(assigned_officers__pk__in=officers) | Q(assignments__officer__pk__in=officers)
            ).distinct()

        qs = qs.select_related('organization', 'created_by').prefetch_related(
            Prefetch(
                'assigned_officers',
                queryset=User.objects.select_related('officer_profile__position')
            )
        )

        events = []
        for task in qs:
            task_date = task.due_date.isoformat() if task.due_date else (task.created_at.date().isoformat() if task.created_at else None)
            if not task_date:
                continue

            event_class = 'fc-event-primary'
            if task.status == 'completed':
                event_class = 'fc-event-success'
            elif task.priority == 'urgent':
                event_class = 'fc-event-danger'
            elif task.priority == 'high':
                event_class = 'fc-event-warning'
            elif task.priority == 'medium':
                event_class = 'fc-event-primary'
            elif task.priority == 'low':
                event_class = 'fc-event-secondary'

            officers = [
                {
                    'id': u.id,
                    'name': u.get_full_name() or u.username,
                    'position': u.position_title,
                    'initials': u.position_initials,
                    'pic_url': u.profile_picture.url if u.profile_picture else None
                } for u in task.assigned_officers.all()
            ]

            events.append({
                'id': task.id,
                'title': task.title,
                'start': task_date,
                'allDay': True,
                'className': event_class,
                'extendedProps': {
                    'task_id': task.id,
                    'task_number': task.task_number,
                    'title': task.title,
                    'description': task.description or 'No description provided.',
                    'category': task.category,
                    'category_display': task.get_category_display(),
                    'category_icon': task.category_icon,
                    'status': task.status,
                    'status_display': task.get_status_display(),
                    'priority': task.priority,
                    'priority_display': task.get_priority_display(),
                    'progress': task.progress,
                    'due_date': task.due_date.strftime('%B %d, %Y') if task.due_date else 'No Due Date',
                    'org_name': task.organization.name if task.organization else 'N/A',
                    'created_by': task.created_by.get_full_name() if task.created_by else 'System',
                    'officers': officers,
                    'detail_url': f"/tasks/{task.id}/"
                }
            })

        return JsonResponse(events, safe=False)


class TaskMoveStatusView(LoginRequiredMixin, View):

    def post(self, request, pk):
        task = get_object_or_404(Task, pk=pk)
        if not request.user.can_update_task_progress(task):
            return JsonResponse({'error': 'Permission denied. You are not authorized to move this task.'}, status=403)
        new_status = request.POST.get('status')
        password = request.POST.get('password', '')
        valid_statuses = [code for code, label in Task.STATUS_CHOICES]

        if new_status not in valid_statuses:
            return JsonResponse({'error': 'Invalid status'}, status=400)

        if new_status == 'completed' or task.status == 'completed':
            if not check_org_admin_password(request.user, password):
                return JsonResponse({'error': 'Incorrect Organization Admin password. Task status update cancelled.'}, status=400)

        old_status = task.status
        if old_status != new_status:
            task.status = new_status
            if new_status == 'completed':
                if not task.completion_date:
                    task.completion_date = timezone.now().date()
                task.progress = 100
            elif old_status == 'completed' and new_status != 'completed':
                if task.progress == 100:
                    task.progress = 99
            task.save()

            TaskHistory.objects.create(
                task=task,
                changed_by=request.user,
                field_changed='Status',
                old_value=dict(Task.STATUS_CHOICES).get(old_status, old_status),
                new_value=dict(Task.STATUS_CHOICES).get(new_status, new_status),
            )

        invalidate_task_caches(task)
        return JsonResponse({
            'status': 'ok',
            'task_id': task.pk,
            'task_number': task.task_number,
            'new_status': task.status,
            'new_status_display': task.get_status_display(),
        })


class TaskDetailJSONView(LoginRequiredMixin, View):
    def get(self, request, pk):
        task = get_object_or_404(
            Task.objects.select_related('created_by')
            .prefetch_related('assigned_officers', 'attachments', 'comments__author'),
            pk=pk
        )

        assigned = [
            {
                'id': u.pk,
                'name': u.get_full_name() or u.username,
                'initials': f"{u.first_name[:1]}{u.last_name[:1]}".upper() if u.first_name and u.last_name else u.username[:2].upper(),
                'role': u.position_title,
                'position_title': u.position_title,
                'email': u.email or 'No email'
            }
            for u in task.sorted_assigned_officers
        ]

        attachments = [
            {
                'id': att.pk,
                'filename': att.filename,
                'url': f'/tasks/attachments/{att.pk}/view/',
                'view_url': f'/tasks/attachments/{att.pk}/view/',
                'download_url': f'/tasks/attachments/{att.pk}/download/',
                'created_at': att.created_at.strftime('%b %d, %Y'),
                'uploaded_by': att.uploaded_by.get_full_name() or att.uploaded_by.username,
                'uploaded_by_role': att.uploaded_by.get_role_display(),
                'icon_name': att.icon_name,
            }
            for att in task.attachments.all()
        ]

        comments = [
            {
                'id': c.pk,
                'author': c.author.get_full_name() or c.author.username,
                'author_username': c.author.username,
                'author_role': c.author.get_role_display(),
                'author_initials': f"{c.author.first_name[:1]}{c.author.last_name[:1]}".upper() if c.author.first_name and c.author.last_name else c.author.username[:2].upper(),
                'content': c.content,
                'created_at': c.created_at.strftime('%b %d, %Y %I:%M %p')
            }
            for c in task.comments.all()
        ]

        data = {
            'id': task.pk,
            'task_number': task.task_number,
            'title': task.title,
            'description': task.description,
            'category': task.category,
            'category_display': task.get_category_display(),
            'category_icon': task.category_icon,
            'status': task.status,
            'status_display': task.get_status_display(),
            'priority': task.priority,
            'priority_display': task.get_priority_display(),
            'progress': task.progress,
            'due_date': task.due_date.strftime('%b %d, %Y') if task.due_date else None,
            'is_overdue': task.is_overdue,
            'created_by': task.created_by.get_full_name() or task.created_by.username,
            'created_at': task.created_at.strftime('%b %d, %Y'),
            'assigned_officers': assigned,
            'attachments': attachments,
            'comments': comments,
            'detail_url': f'/tasks/{task.pk}/',
            'edit_url': f'/tasks/{task.pk}/edit/' if request.user.can_edit_task(task) else None,
            'can_edit': request.user.can_edit_task(task),
            'can_update_progress': request.user.can_update_task_progress(task),
            'can_nudge': request.user.can_nudge_task(task),
            'nudge_url': f'/tasks/{task.pk}/nudge/',
        }

        return JsonResponse(data)


class UpdateProgressView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = Task.objects.filter(pk=pk).first()
        if not task:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')
        if not request.user.can_update_task_progress(task):
            messages.error(request, 'Permission denied. You are not authorized to update progress for this task.')
            return redirect('tasks:detail', pk=pk)
        progress = int(request.POST.get('progress', task.progress))
        status = request.POST.get('status', task.status)
        
        # Link completed status <-> 100% progress
        if status == 'completed' and task.status != 'completed':
            progress = 100
        elif progress == 100 and task.progress != 100:
            status = 'completed'
        elif status != 'completed' and task.status == 'completed' and progress == 100:
            # If changing out of completed but didn't touch progress slider, bump it down slightly
            progress = 99
        old_progress = task.progress
        old_status = task.status
        task.progress = progress
        task.status = status
        if status == 'completed' and not task.completion_date:
            task.completion_date = timezone.now().date()
        task.save()
        if old_progress != progress:
            TaskHistory.objects.create(
                task=task, changed_by=request.user,
                field_changed='Progress', old_value=str(old_progress), new_value=str(progress)
            )
        if old_status != status:
            TaskHistory.objects.create(
                task=task, changed_by=request.user,
                field_changed='Status', old_value=old_status, new_value=status
            )
            # Notify if waiting approval
            if status == 'waiting_approval':
                for admin_user in User.objects.filter(role__in=['super_admin', 'president']):
                    Notification.objects.create(
                        recipient=admin_user,
                        title='Task Waiting Approval',
                        message=f'Task "{task.title}" is waiting for approval.',
                        notification_type='approval_needed',
                        related_task=task
                    )
        messages.success(request, 'Task progress updated.')
        return redirect('tasks:detail', pk=pk)


class MarkCompleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = Task.objects.filter(pk=pk).first()
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('is_ajax') == 'true'

        if not task:
            if is_ajax:
                return JsonResponse({'error': 'This task has already been deleted or no longer exists.'}, status=404)
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')

        if not request.user.can_update_task_progress(task):
            if is_ajax:
                return JsonResponse({'error': 'Permission denied. You are not authorized to mark this task complete.'}, status=403)
            messages.error(request, 'Permission denied. You are not authorized to mark this task complete.')
            return redirect('tasks:detail', pk=pk)

        password = request.POST.get('password', '')
        if not check_org_admin_password(request.user, password):
            if is_ajax:
                return JsonResponse({'error': 'Incorrect admin password. Password verification is required to complete task.'}, status=400)
            messages.error(request, 'Incorrect admin password. Task completion cancelled.')
            next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('tasks:detail', kwargs={'pk': pk})
            return redirect(next_url)

        old_status = task.status
        if old_status != 'completed':
            task.status = 'completed'
            task.completion_date = timezone.now().date()
            task.progress = 100
            task.save()

            TaskHistory.objects.create(
                task=task,
                changed_by=request.user,
                field_changed='Status',
                old_value=dict(Task.STATUS_CHOICES).get(old_status, old_status),
                new_value='Completed'
            )
            for assignment in task.assignments.select_related('officer').all():
                Notification.objects.create(
                    recipient=assignment.officer,
                    title='Task Completed',
                    message=f'Task "{task.title}" has been marked as completed.',
                    notification_type='task_completed',
                    related_task=task
                )

        if is_ajax:
            return JsonResponse({
                'status': 'ok',
                'task_id': task.pk,
                'task_number': task.task_number,
                'new_status': task.status,
                'new_status_display': task.get_status_display(),
                'progress': 100
            })

        messages.success(request, f'Task {task.task_number} marked as completed.')
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('tasks:detail', kwargs={'pk': pk})
        return redirect(next_url)


class ArchiveTaskView(LoginRequiredMixin, View):
    def post(self, request, pk):
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER', '')
        if not next_url or '/tasks/' not in next_url:
            next_url = reverse('tasks:list')

        if not request.user.has_task_override:
            messages.error(request, 'Permission denied. Only Super Admin can archive tasks.')
            return redirect(next_url)
        task = Task.objects.filter(pk=pk).first()
        if not task:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect(next_url)
        task.is_archived = True
        task.save()
        messages.success(request, f'Task {task.task_number} archived.')
        return redirect(next_url)


class AddCommentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = Task.objects.filter(pk=pk).first()
        if not task:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.task = task
            comment.author = request.user
            comment.save()

            # Notify assigned officers & creator (excluding commenter)
            recipients = set(task.assigned_officers.filter(is_active=True))
            if task.created_by and task.created_by.is_active:
                recipients.add(task.created_by)
            recipients.discard(request.user)

            author_name = request.user.get_full_name() or request.user.username
            comment_snippet = comment.content[:80] + ('...' if len(comment.content) > 80 else '')
            for recipient in recipients:
                Notification.objects.create(
                    recipient=recipient,
                    title=f"New Comment on Task {task.task_number}",
                    message=f"{author_name} commented: \"{comment_snippet}\"",
                    notification_type='comment_added',
                    related_task=task
                )

            messages.success(request, 'Comment added.')
        return redirect('tasks:detail', pk=pk)


class DeleteCommentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        comment = TaskComment.objects.filter(pk=pk).first()
        if not comment:
            messages.warning(request, 'This comment has already been deleted or no longer exists.')
            return redirect('tasks:list')
        task_pk = comment.task.pk
        if not (request.user == comment.author or request.user == comment.task.created_by or request.user.has_task_override or request.user.can_edit_task(comment.task)):
            messages.error(request, 'You do not have permission to delete this comment.')
            return redirect('tasks:detail', pk=task_pk)

        comment.delete()
        messages.success(request, 'Comment deleted successfully.')
        return redirect('tasks:detail', pk=task_pk)


class AddAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = Task.objects.filter(pk=pk).first()
        if not task:
            messages.warning(request, 'This task has already been deleted or no longer exists.')
            return redirect('tasks:list')
        form = AttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            attachment = form.save(commit=False)
            attachment.task = task
            attachment.uploaded_by = request.user
            filename = request.FILES['file'].name
            attachment.filename = filename
            attachment.save()

            # Notify assigned officers & creator (excluding uploader)
            recipients = set(task.assigned_officers.filter(is_active=True))
            if task.created_by and task.created_by.is_active:
                recipients.add(task.created_by)
            recipients.discard(request.user)

            uploader_name = request.user.get_full_name() or request.user.username
            for recipient in recipients:
                Notification.objects.create(
                    recipient=recipient,
                    title=f"New Attachment on Task {task.task_number}",
                    message=f"{uploader_name} attached a file: \"{filename}\"",
                    notification_type='attachment_added',
                    related_task=task
                )

            messages.success(request, 'Attachment uploaded.')
        return redirect('tasks:detail', pk=pk)


class DeleteAttachmentView(LoginRequiredMixin, View):
    def post(self, request, pk):
        attachment = TaskAttachment.objects.filter(pk=pk).first()
        if not attachment:
            messages.warning(request, 'This attachment has already been deleted or no longer exists.')
            return redirect('tasks:list')
        task_pk = attachment.task.pk
        attachment.delete()
        messages.success(request, 'Attachment deleted.')
        return redirect('tasks:detail', pk=task_pk)


class DownloadAttachmentView(LoginRequiredMixin, View):
    """Secure attachment downloader forcing direct device file download."""
    def get(self, request, pk):
        import os, logging, urllib.parse, mimetypes, requests
        logger = logging.getLogger(__name__)
        attachment = get_object_or_404(TaskAttachment, pk=pk)

        filename = attachment.filename or os.path.basename(attachment.file.name)
        safe_filename = urllib.parse.quote(filename)
        content_disposition = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{safe_filename}'

        # 1. Try local media storage backup first
        clean_name = attachment.file.name.replace('media/', '', 1) if attachment.file.name.startswith('media/') else attachment.file.name
        full_local_path = os.path.join(django_settings.MEDIA_ROOT, attachment.file.name)
        clean_local_path = os.path.join(django_settings.MEDIA_ROOT, clean_name)

        target_path = None
        if os.path.exists(full_local_path):
            target_path = full_local_path
        elif os.path.exists(clean_local_path):
            target_path = clean_local_path

        if target_path and os.path.isfile(target_path):
            try:
                return FileResponse(open(target_path, 'rb'), content_type='application/octet-stream', as_attachment=True, filename=filename)
            except Exception as e:
                logger.warning(f"Failed serving local attachment file: {e}")

        # 2. Try Django Storage API (att.file.open / read binary bytes)
        content = None
        try:
            attachment.file.open('rb')
            content = attachment.file.read()
            attachment.file.close()

            if content and (content.startswith(b'<!DOCTYPE html') or content.startswith(b'<html') or b'401 Unauthorized' in content[:500]):
                content = None
        except Exception as ex:
            logger.warning(f"Django storage open failed for {attachment.file.name}: {ex}")

        # 3. Try Cloudinary API resource lookup if content missing
        if not content:
            try:
                import cloudinary, cloudinary.api
                pub_id = clean_name
                for r_type in ['raw', 'image']:
                    try:
                        res_info = cloudinary.api.resource(pub_id, resource_type=r_type)
                        sec_url = res_info.get('secure_url')
                        if sec_url:
                            resp = requests.get(sec_url, timeout=10)
                            if resp.status_code == 200 and resp.content:
                                content = resp.content
                                break
                    except Exception:
                        pass
            except Exception as ex:
                logger.warning(f"Cloudinary resource lookup failed: {ex}")

        if content:
            # Cache locally for future instant access
            try:
                cache_path = full_local_path if full_local_path else clean_local_path
                os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                with open(cache_path, 'wb') as cf:
                    cf.write(content)
            except Exception:
                pass

            response = HttpResponse(content, content_type='application/octet-stream')
            response['Content-Disposition'] = content_disposition
            return response

        messages.error(request, f"Unable to download attachment '{filename}'. The file may be unavailable.")
        return redirect('tasks:detail', pk=attachment.task.pk)


class ViewAttachmentView(LoginRequiredMixin, View):
    """View/preview attachment directly in browser with inline Content-Disposition."""
    def get(self, request, pk):
        import os, logging, urllib.parse, mimetypes, requests
        logger = logging.getLogger(__name__)
        attachment = get_object_or_404(TaskAttachment, pk=pk)

        filename = attachment.filename or os.path.basename(attachment.file.name)
        safe_filename = urllib.parse.quote(filename)
        content_disposition = f'inline; filename="{filename}"; filename*=UTF-8\'\'{safe_filename}'

        guessed_type, _ = mimetypes.guess_type(filename)
        content_type = guessed_type or 'application/pdf'

        # 1. Try local file system first
        clean_name = attachment.file.name.replace('media/', '', 1) if attachment.file.name.startswith('media/') else attachment.file.name
        full_local_path = os.path.join(django_settings.MEDIA_ROOT, attachment.file.name)
        clean_local_path = os.path.join(django_settings.MEDIA_ROOT, clean_name)

        target_path = None
        if os.path.exists(full_local_path):
            target_path = full_local_path
        elif os.path.exists(clean_local_path):
            target_path = clean_local_path

        if target_path and os.path.isfile(target_path):
            try:
                response = FileResponse(open(target_path, 'rb'), content_type=content_type)
                response['Content-Disposition'] = content_disposition
                return response
            except Exception as e:
                logger.warning(f"Failed viewing local attachment file: {e}")

        # 2. Try Django Storage API (open/read binary bytes)
        content = None
        try:
            attachment.file.open('rb')
            content = attachment.file.read()
            attachment.file.close()

            if content and (content.startswith(b'<!DOCTYPE html') or content.startswith(b'<html') or b'401 Unauthorized' in content[:500]):
                content = None
        except Exception as ex:
            logger.warning(f"Django storage open failed for {attachment.file.name}: {ex}")

        # 3. Try Cloudinary API resource lookup if content missing
        if not content:
            try:
                import cloudinary, cloudinary.api
                pub_id = clean_name
                for r_type in ['raw', 'image']:
                    try:
                        res_info = cloudinary.api.resource(pub_id, resource_type=r_type)
                        sec_url = res_info.get('secure_url')
                        if sec_url:
                            resp = requests.get(sec_url, timeout=10)
                            if resp.status_code == 200 and resp.content:
                                content = resp.content
                                break
                    except Exception:
                        pass
            except Exception as ex:
                logger.warning(f"Cloudinary resource lookup failed: {ex}")

        if content:
            # Cache locally for future instant access
            try:
                cache_path = full_local_path if full_local_path else clean_local_path
                os.makedirs(os.path.dirname(cache_path), exist_ok=True)
                with open(cache_path, 'wb') as cf:
                    cf.write(content)
            except Exception:
                pass

            response = HttpResponse(content, content_type=content_type)
            response['Content-Disposition'] = content_disposition
            return response

        messages.error(request, f"Unable to view attachment '{filename}'. The file may be unavailable.")
        return redirect('tasks:detail', pk=attachment.task.pk)



def _get_filtered_tasks_for_export(request):
    if request.user.organization:
        qs = Task.objects.filter(is_archived=False, organization=request.user.organization)
    else:
        qs = Task.objects.filter(is_archived=False)

    scope = request.GET.get('scope', 'all' if request.user.has_task_override else 'my_tasks')
    if scope == 'my_tasks':
        qs = qs.filter(Q(assigned_officers=request.user) | Q(created_by=request.user)).distinct()

    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(task_number__icontains=q) | Q(description__icontains=q))

    status = request.GET.get('status', '')
    if status:
        if status == 'active':
            qs = qs.exclude(status='completed')
        elif status == 'overdue':
            qs = qs.filter(due_date__lt=timezone.now().date()).exclude(status='completed')
        elif status == 'in_progress':
            qs = qs.exclude(status__in=['not_started', 'completed'])
        else:
            qs = qs.filter(status=status)

    priority = request.GET.get('priority', '')
    if priority:
        qs = qs.filter(priority=priority)

    officers = request.GET.getlist('officer')
    if officers:
        qs = qs.filter(assigned_officers__id__in=officers).distinct()

    return qs.select_related('created_by').prefetch_related('assigned_officers')


class ExportTasksPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        tasks = _get_filtered_tasks_for_export(request)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, spaceAfter=6, textColor=colors.HexColor('#1e3a5f'))
        story.append(Paragraph('CSG Task Management Report', title_style))
        story.append(Paragraph(f'Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}', styles['Normal']))
        story.append(Spacer(1, 0.2*inch))

        data = [['Task No.', 'Title', 'Status', 'Priority', 'Due Date', 'Progress']]
        # Use iterator with chunked fetching for large querysets to limit peak memory
        task_count = tasks.count()
        task_iter = tasks.iterator(chunk_size=200) if task_count > 500 else tasks
        for t in task_iter:
            data.append([
                t.task_number,
                t.title[:45],
                t.get_status_display(),
                t.get_priority_display(),
                str(t.due_date) if t.due_date else 'N/A',
                f'{t.progress}%'
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="csg_tasks_report.pdf"'
        return response


class ExportTasksExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        tasks = _get_filtered_tasks_for_export(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tasks Report'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')

        headers = ['Task Number', 'Title', 'Description', 'Status', 'Priority', 'Assigned Officers', 'Due Date', 'Completion Date', 'Progress (%)', 'Created By', 'Created At']
        ws.append(headers)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Use iterator with chunked fetching for large querysets to limit peak memory
        task_count = tasks.count()
        task_iter = tasks.iterator(chunk_size=200) if task_count > 500 else tasks
        for i, t in enumerate(task_iter, 2):
            officers = ', '.join([o.get_full_name() or o.username for o in t.assigned_officers.all()])
            ws.append([
                t.task_number,
                t.title,
                t.description[:100],
                t.get_status_display(),
                t.get_priority_display(),
                officers,
                str(t.due_date) if t.due_date else '',
                str(t.completion_date) if t.completion_date else '',
                t.progress,
                t.created_by.get_full_name() or t.created_by.username,
                t.created_at.strftime('%Y-%m-%d %H:%M'),
            ])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = alt_fill

        col_widths = [15, 35, 40, 15, 12, 30, 12, 15, 12, 20, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = 'A2'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="csg_tasks_report.xlsx"'
        return response


class NudgeOfficersView(LoginRequiredMixin, View):
    """Send a nudge reminder (in-app notification + email) to selected assigned officers."""

    def post(self, request, pk):
        task = get_object_or_404(Task, pk=pk)
        if not request.user.can_nudge_task(task):
            return JsonResponse({'error': 'Permission denied. You are not authorized to nudge this task.'}, status=403)

        # Get the list of officer IDs selected by the sender
        officer_ids = request.POST.getlist('officer_ids')
        if not officer_ids:
            return JsonResponse({'error': 'No officers selected.'}, status=400)

        # Only nudge officers who are actually assigned to this task
        officers = task.assigned_officers.filter(pk__in=officer_ids)
        if not officers.exists():
            return JsonResponse({'error': 'No valid officers found.'}, status=400)

        custom_message = request.POST.get('message', '').strip()
        due_str = task.due_date.strftime('%B %d, %Y') if task.due_date else 'No due date set'
        sender_name = request.user.get_full_name() or request.user.username

        nudged = []
        failed_emails = []
        no_email_officers = []

        import logging
        logger = logging.getLogger(__name__)

        for officer in officers:
            # ── In-app notification ──────────────────────────────
            notif_message = (
                f'{sender_name} sent you a nudge for task "{task.title}" '
                f'[{task.task_number}]. Current progress: {task.progress}%. '
                f'Due: {due_str}.'
            )
            if custom_message:
                notif_message += f' Note: {custom_message}'

            Notification.objects.create(
                recipient=officer,
                title=f'\U0001f514 Nudge: {task.task_number}',
                message=notif_message,
                notification_type='task_updated',
                related_task=task,
            )

            # ── Email ─────────────────────────────────────────────
            if officer.email and '@' in officer.email:
                subject = f'[CSG] Nudge Reminder – {task.task_number}: {task.title}'
                body = (
                    f'Hi {officer.get_full_name() or officer.username},\n\n'
                    f'{sender_name} sent you a nudge for the following task:\n\n'
                    f'  Task No.:  {task.task_number}\n'
                    f'  Title:     {task.title}\n'
                    f'  Status:    {task.get_status_display()}\n'
                    f'  Progress:  {task.progress}%\n'
                    f'  Due Date:  {due_str}\n'
                )
                if custom_message:
                    body += f'\nMessage from {sender_name}:\n  "{custom_message}"\n'
                body += (
                    f'\nPlease log in to the CSG Task Management System to view and '
                    f'update this task.\n\n'
                    f'— CSG Task Management System'
                )
                
                # Email Dispatchers (Brevo HTTPS API -> Resend HTTPS API -> Django SMTP)
                brevo_key = config('BREVO_API_KEY', default='').strip()
                resend_key = config('RESEND_API_KEY', default='').strip()
                email_sent = False
                email_err = None

                # 1. Try Brevo HTTPS API (300 free emails/day to ANY recipient address)
                if brevo_key:
                    try:
                        import requests
                        resp = requests.post(
                            "https://api.brevo.com/v3/smtp/email",
                            headers={
                                "api-key": brevo_key,
                                "Content-Type": "application/json",
                                "Accept": "application/json"
                            },
                            json={
                                "sender": {
                                    "name": "CSG System",
                                    "email": config('BREVO_SENDER_EMAIL', default=config('DEFAULT_FROM_EMAIL', default='csgtasks2026@gmail.com')).strip()
                                },
                                "to": [{"email": officer.email}],
                                "subject": subject,
                                "textContent": body,
                            },
                            timeout=10
                        )
                        if resp.status_code in [200, 201]:
                            email_sent = True
                        else:
                            email_err = f"Brevo API ({resp.status_code}): {resp.text}"
                    except Exception as b_ex:
                        email_err = f"Brevo HTTPS failed: {b_ex}"

                # 2. Try Resend HTTPS API (fallback)
                if not email_sent and resend_key:
                    try:
                        import requests
                        resp = requests.post(
                            "https://api.resend.com/emails",
                            headers={
                                "Authorization": f"Bearer {resend_key}",
                                "Content-Type": "application/json"
                            },
                            json={
                                "from": config('RESEND_FROM_EMAIL', default="CSG Task System <onboarding@resend.dev>"),
                                "to": [officer.email],
                                "subject": subject,
                                "text": body,
                            },
                            timeout=10
                        )
                        if resp.status_code in [200, 201]:
                            email_sent = True
                        else:
                            email_err = f"Resend API ({resp.status_code}): {resp.text}"
                    except Exception as r_ex:
                        email_err = f"Resend HTTPS failed: {r_ex}"

                # 3. Try Standard SMTP (if host credentials configured)
                if not email_sent and not brevo_key and not resend_key:
                    if getattr(django_settings, 'EMAIL_HOST_USER', '') and getattr(django_settings, 'EMAIL_HOST_PASSWORD', ''):
                        try:
                            from_email = django_settings.DEFAULT_FROM_EMAIL or django_settings.EMAIL_HOST_USER
                            send_mail(
                                subject=subject,
                                message=body,
                                from_email=from_email,
                                recipient_list=[officer.email],
                                fail_silently=False,
                            )
                            email_sent = True
                        except Exception as e:
                            logger.warning(f"Nudge SMTP email failed for {officer.email}: {e}")
                            email_err = "SMTP delivery unavailable"
                    else:
                        logger.info(f"SMTP credentials not configured for sending email to {officer.email}")

                if not email_sent and email_err:
                    logger.warning(f"Nudge email notice for {officer.email}: {email_err}")
                    failed_emails.append(officer.email)
            else:
                no_email_officers.append(officer.get_full_name() or officer.username)

            nudged.append(officer.get_full_name() or officer.username)

        # Log in task history
        TaskHistory.objects.create(
            task=task,
            changed_by=request.user,
            field_changed='Nudge',
            old_value='',
            new_value=f'Nudge sent to: {", ".join(nudged)}',
        )

        msg = f'In-app nudge sent to {len(nudged)} officer(s) successfully.'
        if no_email_officers:
            msg += f' (No email address configured for: {", ".join(no_email_officers)}).'

        return JsonResponse({
            'ok': True,
            'nudged': nudged,
            'failed_emails': failed_emails,
            'message': msg,
        })


class TaskSearchSuggestionsView(LoginRequiredMixin, View):
    """Returns matching tasks for live search suggestions."""
    def get(self, request, *args, **kwargs):
        q = request.GET.get('q', '').strip()
        if not q:
            return JsonResponse({'results': []})

        qs = Task.objects.filter(is_archived=False)
        if request.user.organization:
            qs = qs.filter(organization=request.user.organization)
        
        qs = qs.filter(
            Q(title__icontains=q) | Q(task_number__icontains=q) | Q(description__icontains=q)
        ).distinct()[:8]

        results = []
        for t in qs:
            results.append({
                'id': t.pk,
                'task_number': t.task_number,
                'title': t.title,
                'status': t.status,
                'status_display': t.get_status_display(),
                'priority': t.priority,
                'due_date': t.due_date.strftime('%b %d, %Y') if t.due_date else '',
                'progress': t.progress,
            })

        return JsonResponse({'results': results})

