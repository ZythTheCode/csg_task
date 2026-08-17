from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic import TemplateView, View
from django.http import HttpResponse
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Count, Q
from core.mixins import FragmentResponseMixin
from core.query_utils import get_report_counts
from tasks.models import Task
from officers.models import Officer
from accounts.models import User
import io
import datetime


class ReportsDashboardView(FragmentResponseMixin, LoginRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    paginate_by = 25

    def dispatch(self, request, *args, **kwargs):
        if not request.user.can_view_reports:
            from django.shortcuts import redirect
            from django.contrib import messages
            messages.error(request, 'You do not have permission to view reports.')
            return redirect('core:dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Reports'
        ctx['status_choices'] = Task.STATUS_CHOICES
        ctx['priority_choices'] = Task.PRIORITY_CHOICES
        org = self.request.user.get_organization(self.request)
        
        # Cache officers list for filter dropdown (TTL 300s)
        org_id = org.pk if org else 'all'
        cache_key = f'reports_officers_{org_id}'
        officers = cache.get(cache_key)
        if officers is None:
            officers_qs = Officer.objects.select_related('user').exclude(user__role__in=['super_admin', 'super_super_admin'])
            if org:
                officers_qs = officers_qs.filter(user__organization=org)
            officers = list(officers_qs.all())
            cache.set(cache_key, officers, 300)
        ctx['officers'] = officers

        # Apply filters
        filters = self._get_filters()
        tasks = self._get_filtered_tasks(filters)
        ctx['filters'] = filters
        ctx['scope'] = filters['scope']

        # Use single aggregate query for all counts instead of multiple count() calls
        today = timezone.now().date()
        report_counts = get_report_counts(tasks, today)
        ctx['task_count'] = report_counts['total']
        ctx['active_count'] = report_counts['active']
        ctx['completed_count'] = report_counts['completed']

        # Summary stats from the single aggregate (no additional queries)
        ctx['summary'] = {
            'total': report_counts['total'],
            'completed': report_counts['completed'],
            'overdue': report_counts['overdue'],
            'in_progress': report_counts['in_progress'],
        }

        # Paginate active tasks at 25 items per page
        active_tasks_qs = tasks.exclude(status='completed')
        active_paginator = Paginator(active_tasks_qs, self.paginate_by)
        active_page_number = self.request.GET.get('active_page', 1)
        try:
            active_page = active_paginator.page(active_page_number)
        except PageNotAnInteger:
            active_page = active_paginator.page(1)
        except EmptyPage:
            active_page = active_paginator.page(active_paginator.num_pages)

        ctx['active_tasks'] = active_page
        ctx['active_page_obj'] = active_page

        # Paginate completed tasks at 25 items per page
        completed_tasks_qs = tasks.filter(status='completed')
        completed_paginator = Paginator(completed_tasks_qs, self.paginate_by)
        completed_page_number = self.request.GET.get('completed_page', 1)
        try:
            completed_page = completed_paginator.page(completed_page_number)
        except PageNotAnInteger:
            completed_page = completed_paginator.page(1)
        except EmptyPage:
            completed_page = completed_paginator.page(completed_paginator.num_pages)

        ctx['completed_tasks'] = completed_page
        ctx['completed_page_obj'] = completed_page

        # Keep full queryset reference for exports (not evaluated here)
        ctx['tasks'] = tasks
        return ctx

    def _get_filters(self):
        req = self.request
        return {
            'officer': req.GET.get('officer', ''),
            'month': req.GET.get('month', ''),
            'year': req.GET.get('year', ''),
            'status': req.GET.get('status', ''),
            'priority': req.GET.get('priority', ''),
            'scope': req.GET.get('scope', 'all' if req.user.has_task_override else 'my_tasks'),
        }

    def _get_filtered_tasks(self, filters):
        qs = Task.objects.filter(is_archived=False).select_related('created_by', 'organization').prefetch_related('assigned_officers', 'assigned_officers__officer_profile', 'assigned_officers__officer_profile__position')
        org = self.request.user.get_organization(self.request)
        if org:
            qs = qs.filter(organization=org)
            
        if filters['scope'] == 'my_tasks':
            qs = qs.filter(Q(assigned_officers=self.request.user) | Q(created_by=self.request.user)).distinct()

        if filters['officer']:
            qs = qs.filter(assigned_officers__id=filters['officer'])
        if filters['year']:
            try:
                y = int(filters['year'])
                qs = qs.filter(Q(due_date__year=y) | Q(created_at__year=y))
            except ValueError:
                pass
        if filters['month']:
            try:
                m = int(filters['month'])
                qs = qs.filter(Q(due_date__month=m) | Q(created_at__month=m))
            except ValueError:
                pass
        if filters['status']:
            if filters['status'] == 'in_progress':
                qs = qs.exclude(status__in=['not_started', 'completed'])
            else:
                qs = qs.filter(status=filters['status'])
        if filters['priority']:
            qs = qs.filter(priority=filters['priority'])
        return qs


class ExportReportPDFView(LoginRequiredMixin, View):
    def get(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        officer_id = request.GET.get('officer', '')
        year = request.GET.get('year', str(timezone.now().year))
        month = request.GET.get('month', '')
        status = request.GET.get('status', '')
        priority = request.GET.get('priority', '')
        task_ids = request.GET.get('task_ids', '')

        scope = request.GET.get('scope', 'all' if request.user.has_task_override else 'my_tasks')

        qs = Task.objects.filter(is_archived=False).select_related('created_by')
        if request.user.organization:
            qs = qs.filter(organization=request.user.organization)
            
        if task_ids:
            # If task_ids are provided, we only export those specific tasks
            qs = qs.filter(id__in=task_ids.split(','))
        else:
            if scope == 'my_tasks':
                from django.db.models import Q
                qs = qs.filter(Q(assigned_officers=request.user) | Q(created_by=request.user)).distinct()
            # Otherwise apply the normal filters
            if officer_id:
                qs = qs.filter(assigned_officers__id=officer_id)
            if year:
                qs = qs.filter(created_at__year=year)
            if month:
                qs = qs.filter(created_at__month=month)
            if status:
                if status == 'in_progress':
                    qs = qs.exclude(status__in=['not_started', 'completed'])
                else:
                    qs = qs.filter(status=status)
            if priority:
                qs = qs.filter(priority=priority)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=20, spaceAfter=4, textColor=colors.HexColor('#1e3a5f'))
        sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#666666'))

        task_count = qs.count()
        story.append(Paragraph('CSG Task Management Report', title_style))
        story.append(Paragraph(f'Generated: {timezone.now().strftime("%B %d, %Y %I:%M %p")}  |  Total Tasks: {task_count}', sub_style))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1e3a5f'), spaceAfter=10))

        data = [['Task No.', 'Title', 'Status', 'Priority', 'Assigned To', 'Due Date', 'Progress']]
        # Use iterator with chunked fetching for large querysets to limit peak memory.
        # For large exports (>500), drop prefetch_related since iterator() ignores it;
        # the N+1 on assigned_officers is acceptable for batch file exports.
        if task_count > 500:
            task_iter = qs.iterator(chunk_size=200)
        else:
            task_iter = qs.prefetch_related('assigned_officers')
        for t in task_iter:
            officers = ', '.join([f"{o.get_full_name() or o.username} ({o.position_initials})" for o in t.sorted_assigned_officers])
            data.append([
                t.task_number,
                Paragraph(t.title[:45], styles['Normal']),
                t.get_status_display(),
                t.get_priority_display(),
                Paragraph(officers[:35], styles['Normal']),
                str(t.due_date) if t.due_date else 'N/A',
                f'{t.progress}%'
            ])

        col_widths = [1.2*inch, 2.8*inch, 1.2*inch, 1.0*inch, 2.2*inch, 1.2*inch, 0.9*inch]
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="csg_report.pdf"'
        return response


class ExportReportExcelView(LoginRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        officer_id = request.GET.get('officer', '')
        year = request.GET.get('year', str(timezone.now().year))
        month = request.GET.get('month', '')
        status = request.GET.get('status', '')
        priority = request.GET.get('priority', '')
        task_ids = request.GET.get('task_ids', '')

        scope = request.GET.get('scope', 'all' if request.user.has_task_override else 'my_tasks')

        qs = Task.objects.filter(is_archived=False).select_related('created_by')
        if request.user.organization:
            qs = qs.filter(organization=request.user.organization)
            
        if task_ids:
            # If task_ids are provided, we only export those specific tasks
            qs = qs.filter(id__in=task_ids.split(','))
        else:
            if scope == 'my_tasks':
                from django.db.models import Q
                qs = qs.filter(Q(assigned_officers=request.user) | Q(created_by=request.user)).distinct()
            # Otherwise apply the normal filters
            if officer_id:
                qs = qs.filter(assigned_officers__id=officer_id)
            if year:
                qs = qs.filter(created_at__year=year)
            if month:
                qs = qs.filter(created_at__month=month)
            if status:
                if status == 'in_progress':
                    qs = qs.exclude(status__in=['not_started', 'completed'])
                else:
                    qs = qs.filter(status=status)
            if priority:
                qs = qs.filter(priority=priority)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CSG Report'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        alt_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')

        headers = ['Task Number', 'Title', 'Status', 'Priority', 'Assigned Officers', 'Due Date', 'Completion Date', 'Progress (%)', 'Created By']
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Use iterator with chunked fetching for large querysets to limit peak memory.
        # For large exports (>500), drop prefetch_related since iterator() ignores it;
        # the N+1 on assigned_officers is acceptable for batch file exports.
        task_count = qs.count()
        if task_count > 500:
            task_iter = qs.iterator(chunk_size=200)
        else:
            task_iter = qs.prefetch_related('assigned_officers')
        for i, t in enumerate(task_iter, 2):
            officers = ', '.join([f"{o.get_full_name() or o.username} ({o.position_initials})" for o in t.sorted_assigned_officers])
            ws.append([
                t.task_number, t.title, t.get_status_display(), t.get_priority_display(),
                officers,
                str(t.due_date) if t.due_date else '',
                str(t.completion_date) if t.completion_date else '',
                t.progress,
                t.created_by.get_full_name() or t.created_by.username,
            ])
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=i, column=col).fill = alt_fill

        widths = [15, 35, 15, 12, 35, 12, 15, 12, 20]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="csg_report.xlsx"'
        return response
